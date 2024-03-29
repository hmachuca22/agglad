# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook

from django.contrib import messages
from django.db.models import Count, Q
from django.urls import reverse, reverse_lazy
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponseRedirect, HttpResponse
from django.views.generic import CreateView, TemplateView, View
from django.views.generic.edit import FormView

from training.core.models import State, County, Tag
from training.organizations.models import Organization, UserOrganization
from .forms import UserCreationForm, UserForm, UserAcademicDegreeFormSet, UserExternalTrainingFormSet
from .models import StudentProfile, UserAcademicDegree, UserExternalTraining, User
from .mixins import RoleRequiredMixin


class UserDashboardView(RoleRequiredMixin, TemplateView):
    http_method_names = ["get"]
    template_name = "users/dashboard.html"
    allowed_roles = [User.Role.ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        stats = User.objects.aggregate(
            male_total=Count("pk", filter=Q(gender=User.Gender.MALE)),
            male_teacher=Count("pk", filter=Q(gender=User.Gender.MALE, is_teacher=True)),
            male_student=Count("pk", filter=Q(gender=User.Gender.MALE, is_student=True)),
            female_total=Count("pk", filter=Q(gender=User.Gender.FEMALE)),
            female_teacher=Count("pk", filter=Q(gender=User.Gender.FEMALE, is_teacher=True)),
            female_student=Count("pk", filter=Q(gender=User.Gender.FEMALE, is_student=True)),
            other_total=Count("pk", filter=Q(gender=User.Gender.OTHER)),
            other_teacher=Count("pk", filter=Q(gender=User.Gender.OTHER, is_teacher=True)),
            other_student=Count("pk", filter=Q(gender=User.Gender.OTHER, is_student=True)),
        )

        context["stats"] = {
            "male": {
                "total": stats["male_total"],
                "teacher": stats["male_teacher"],
                "student": stats["male_student"],
            },
            "female": {
                "total": stats["female_total"],
                "teacher": stats["female_teacher"],
                "student": stats["female_student"],
            },
            "other": {
                "total": stats["other_total"],
                "teacher": stats["other_teacher"],
                "student": stats["other_student"],
            }
        }

        return context


class UserListView(RoleRequiredMixin, TemplateView):
    http_method_names = ["get"]
    template_name = "users/users.html"
    allowed_roles = [User.Role.ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["genders"] = User.Gender.CHOICES
        context["roles"] = self._get_roles()
        return context

    def _get_roles(self):
        roles = dict(User.Role.CHOICES)
        roles.pop(User.Role.SUPERUSER, None)  # Se remueve el tipo de usuario `SUPERUSER`
        return tuple(roles.items())


class UserCUFormView(RoleRequiredMixin, FormView):  # CU stands for Create and Update
    class Action:
        CREATE = "create"
        UPDATE = "update"

    http_method_names = ["get", "post"]
    template_name = "users/user.html"
    success_url = reverse_lazy("users:users")
    form_class = UserForm

    def __init__(self):
        super().__init__()
        self.user = None
        self.user_academic_degree_form_set = None
        self.user_external_training_form_set = None

    def get_allowed_roles(self):
        if self.kwargs.get("action") == self.Action.UPDATE and not self.kwargs.get("pk"):
            return User.Role.ALL
        return [User.Role.ADMIN]

    def dispatch(self, request, *args, **kwargs):
        if self.kwargs.get("action") == self.Action.UPDATE:
            user_pk = self.kwargs.get("pk")
            if user_pk:
                self.user = get_object_or_404(User, pk=user_pk)
            else:
                self.user = self.request.user

        return super().dispatch(request, *args, **kwargs)

    def get_form(self, form_class=None):
        user_form_params = dict()
        user_academic_degree_form_set_params = dict(prefix="academic_degree")
        user_external_training_form_set_params = dict(prefix="training")

        if self.request.method == "GET":
            if self.user:
                user_form_params.update(instance=self.user)
                user_academic_degree_form_set_params.update(queryset=self.user.useracademicdegree_set.all())
                user_external_training_form_set_params.update(queryset=self.user.userexternaltraining_set.all())
            else:
                user_academic_degree_form_set_params.update(queryset=UserAcademicDegree.objects.none())
                user_external_training_form_set_params.update(queryset=UserExternalTraining.objects.none())
        else:
            user_form_params.update(data=self.request.POST)
            user_academic_degree_form_set_params.update(data=self.request.POST)
            user_external_training_form_set_params.update(data=self.request.POST)

            if self.user:
                user_form_params.update(instance=self.user)

        user_form = UserForm(**user_form_params)
        self.user_academic_degree_form_set = UserAcademicDegreeFormSet(**user_academic_degree_form_set_params)
        self.user_external_training_form_set = UserExternalTrainingFormSet(**user_external_training_form_set_params)

        return user_form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["user"] = self.user
        context["user_organization"] = self.get_user_organization()
        context["user_academic_degree_form_set"] = self.user_academic_degree_form_set
        context["user_external_training_form_set"] = self.user_external_training_form_set
        context["states"] = [{
            "id": state.id,
            "name": state.name,
            "counties": [{
                "id": county.id,
                "name": county.name
            } for county in County.objects.filter(parent=state).order_by("code")]
        } for state in State.objects.all().order_by("code")]
        context["organizations"] = self.get_organizations()

        return context

    def get_organizations(self):
        organizations = []
        user_organization = UserOrganization.objects.filter(user=self.user).first()

        for organization_type, organization_label in Organization.Type.CHOICES:
            organizations_list = [{
                "id": instance.id,
                "name": instance.name
            } for instance in Organization.objects.filter(type=organization_type).order_by("code")]
            if organizations_list:
                organizations.append({"type_label": organization_label, "items": organizations_list})

        return organizations

    def get_user_organization(self):
        if self.user and self.user.is_organizational:
            return UserOrganization.objects.filter(user=self.user).first()
        return None

    def form_valid(self, form):
        user = form.save(commit=False)
        user.set_password("training")  # TODO: Corregir
        user.save()

        if user.is_organizational:
            organization = get_object_or_404(Organization, pk=form.cleaned_data["organization"])
            UserOrganization.objects.get_or_create(user=user, organization=organization)
        else:
            UserOrganization.objects.filter(user=user).delete()

        for user_academic_degree_form in self.user_academic_degree_form_set:
            cleaned_data = user_academic_degree_form.cleaned_data
            if cleaned_data:
                user_academic_degree = user_academic_degree_form.save(commit=False)
                if cleaned_data.get("DELETE"):
                    user_academic_degree.delete()
                else:
                    user_academic_degree.user = user
                    user_academic_degree.save()

        for user_external_training_form in self.user_external_training_form_set:
            cleaned_data = user_external_training_form.cleaned_data
            if cleaned_data:
                user_external_training = user_external_training_form.save(commit=False)
                if cleaned_data.get("DELETE"):
                    user_external_training.delete()
                else:
                    user_external_training.user = user
                    user_external_training.save()

                    user_external_training.tags.clear()
                    for tag in cleaned_data.get('tags'):
                        user_external_training.tags.add(tag)

        return HttpResponseRedirect(self.get_success_url())

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if (
            form.is_valid()
            and self.user_academic_degree_form_set.is_valid()
            and self.user_external_training_form_set.is_valid()
        ):
            return self.form_valid(form)
        return self.form_invalid(form)


class UserProfileView(RoleRequiredMixin, TemplateView):
    allowed_roles = User.Role.ALL
    http_method_names = ["get"]
    template_name = "users/profile.html"

    def dispatch(self, request, *args, **kwargs):
        user_pk = kwargs.get("pk")
        if user_pk:
            self.user = get_object_or_404(User, pk=user_pk)
        else:
            self.user = self.request.user
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        academic_records = self.user.get_achievements_record()
        tags = []
        for record in academic_records:
            if record.get('tags'):
                for tag in record.get('tags'):
                    tags.append(tag[0])
        acquired_knowledge = list(set(tags))

        historic_enroll = self.user.get_historic_enroll()

        context['user'] = self.user
        context["academic_degree_types"] = UserAcademicDegree.Type.ALL
        context["external_training_types"] = UserExternalTraining.Type.CHOICES
        context['academic_record'] = academic_records
        context['acquired_knowledge'] = acquired_knowledge
        context['historic_enroll'] = historic_enroll
        return context


class RegisterAccountView(CreateView):
    model = User
    template_name = "pages/register.html"
    form = UserCreationForm
    fields = [
        'username',
        'password',
        'email',
        'first_name',
        'last_name',
        'gender',
        'birth_day',
    ]
    is_bound = True

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            instance = form.save()
            instance.set_password(form.cleaned_data['password'])
            instance.is_student = True
            student = StudentProfile.objects.get_or_create(student=instance)
            instance.save()
        else:
            messages.warning = form.errors.as_json()

        return HttpResponseRedirect(reverse('account_login'))


class UserExternalTrainingDump(View):
    def get(self, request):
        template_name = "users/external_training_dump.html"
        return render(request, template_name)

    def post(self, request, *args, **kwargs):
        type_map = dict(UserExternalTraining.Type.CHOICES)
        type_list = {v: k for k, v in type_map.items()}

        modality_map = dict(UserExternalTraining.Modality.CHOICES)
        modality_list = {v: k for k, v in modality_map.items()}

        try:
            dumpfile = request.FILES.get("dumpfile")
            wb = load_workbook(dumpfile)
            sh = wb.active

            for r in range(1, sh.max_row + 1):
                if r > 1:
                    cell_a = 'A{}'.format(r, )
                    cell_c = 'C{}'.format(r, )
                    cell_d = 'D{}'.format(r, )
                    cell_e = 'E{}'.format(r, )
                    cell_f = 'F{}'.format(r, )
                    cell_g = 'G{}'.format(r, )
                    cell_h = 'H{}'.format(r, )
                    cell_i = 'I{}'.format(r, )
                    cell_j = 'J{}'.format(r, )
                    cell_k = 'K{}'.format(r, )

                    text_user_id = sh[cell_a].value
                    text_name = sh[cell_c].value
                    text_description = sh[cell_d].value
                    text_type = sh[cell_e].value
                    text_modality = sh[cell_f].value
                    text_location = sh[cell_g].value
                    text_started_at = sh[cell_h].value
                    text_finished_at = sh[cell_i].value
                    text_duration = sh[cell_j].value
                    text_tags = sh[cell_k].value

                    try:
                        user_obj = User.objects.get(username=text_user_id)
                    except Exception as e:
                        message = 'El Usuario <b>"{}"</b> (línea <b>{}</b>) no existe, por favor verificar conforme a las instrucciones del archivo e \
                            <a href="{}" class="btn-link"><u>intentar nuevamente</u></a>.'.format(
                            text_user_id,
                            r,
                            reverse('users:add-external-training-dump')
                        )
                        context = {'title': 'Error', 'message': message}
                        return render(request, "users/external_training_result.html", context)

                    type_obj = type_list.get(text_type, "")
                    if type_obj == "":
                        message = 'El Tipo de Curso <b>"{}"</b> (línea <b>{}</b>) no existe, por favor verificar conforme a las instrucciones del archivo e \
                            <a href="{}" class="btn-link"><u>intentar nuevamente</u></a>.'.format(
                            text_type,
                            r,
                            reverse('users:add-external-training-dump')
                        )
                        context = {'title': 'Error', 'message': message}
                        return render(request, "users/external_training_result.html", context)

                    modality_obj = modality_list.get(text_modality, "")
                    if modality_obj == "":
                        message = 'El Tipo de Modalidad <b>"{}"</b> (línea <b>{}</b>) no existe, por favor verificar conforme a las instrucciones del archivo e \
                            <a href="{}" class="btn-link"><u>intentar nuevamente</u></a>.'.format(
                            text_modality,
                            reverse('users:add-external-training-dump')
                        )
                        context = {'title': 'Error', 'message': message}
                        return render(request, "users/external_training_result.html", context)

                    external_training = UserExternalTraining.objects.create(
                        user=user_obj,
                        name=text_name,
                        description=text_description,
                        type=type_obj,
                        modality=modality_obj,
                        location=text_location,
                        started_at=text_started_at,
                        finished_at=text_finished_at,
                        duration=text_duration
                    )
                    external_training.save()


        except Exception as e:
            message = 'Existen datos erróneos en el archivo, por favor verificar conforme a las instrucciones del archivo e \
                <a href="{}" class="btn-link"><u>intentar nuevamente</u></a>.'.format(
                reverse('users:add-external-training-dump'))
            context = {'title': 'Error', 'message': message}
            return render(request, "users/external_training_result.html", context)

        tags=text_tags.split(',')

        if tags:
            for tag in tags:
                try:
                    external_training.tags.add(Tag.objects.get(slug=tag.strip()))
                except Exception as e:
                    pass

        return HttpResponseRedirect(reverse('users:external-training-success'))


class UserExternalTrainingSuccess(TemplateView):
    template_name = "users/external_training_result.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = 'Éxito'
        context[
            "message"] = '¡Datos gardados con éxito! <a href="{}" class="btn-link"><u>Ir a Usuarios</u></a>.'.format(
            reverse('users:users'))
        return context


class UserExternalTrainingTemplate(View):
    def get(self, request):
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('Formato')

        ws.append([
            'Identidad',
            'Nombre Completo',
            'Nombre de Capacitación',
            'Descripción Corta de Capacitación',
            'Tipo',
            'Modalidad',
            'Lugar',
            'Fecha Inicial de Entrenamiento',
            'Fecha Final de Entrenamiento',
            'Duración (Hrs)',
            'Tema'
        ])

        wi = wb.create_sheet('Instrucciones')

        wi.append(['Identidad: Asegurar que sea formato texto, sin guiones ni caracteres especiales', ])
        wi.append(['Nombre Completo: Nombre completo de la persona', ])
        wi.append(
            ['Opciones Tipo de Capacitación: "Capacitación", "Diplomado", "Curso", "Taller", "Seminario", "Módulo" ', ])
        wi.append(['Opciones de Modalidad: "Presencial", "Virtual", "Presencial/Virtual" ', ])
        wi.append(['Fechas: Formato dd/mm/aaaa (numérico)', ])
        wi.append(['Duración (Hrs): indicar solamente el valor entero de horas', ])
        wi.append(['', ])
        wi.append(['No se debe modificar el orden de las columnas', ])
        wi.append(['No se debe eliminar el encabezado de documento', ])
        wi.append(['Se deben respetar las opciones de tipo y modalidad de capacitación', ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=entrenamiento_externo.xlsx'

        wb.save(response)

        return response
